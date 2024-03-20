# pip install mysql-connector-python
# pip install pytz
#pip install openpyxl

import csv
import re
import mysql.connector
from datetime import datetime, timezone, timedelta
import pytz
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime
import pytz

# Diccionario de condiciones
condicion_dict = {'BUENO': 1, 'REGULAR': 0, 'DAÑADO': -1}
estado_dict = {
    'DISPONIBLE': 1,
    'PRESTADO': 2,
    'AGOTADO': 3,
    'DAÑADO': 4,
    'EN REPARACION': 5,
    'RETIRADO': 6,
    'APARTADO': 7
}

def getCategoria(tipo, categoria):
    cursor = connection.cursor()
    
    # Mapea el tipo a un entero
    tipo = 1 if tipo.lower() == 'equipo' else 2

    # Verifica si la categoría ya existe en la base de datos
    query = "SELECT Id FROM Categoria WHERE Tipo = %s AND Nombre = %s"
    cursor.execute(query, (tipo, categoria))
    result = cursor.fetchone()

    if result:
        # Si existe, devuelve el ID de la categoría existente
        return result[0]
    else:
        # Si no existe, crea la categoría y devuelve el ID recién creado
        query = "INSERT INTO Categoria (Tipo, Nombre) VALUES (%s, %s)"
        cursor.execute(query, (tipo, categoria))
        connection.commit()
        return cursor.lastrowid


def getActivoBodega(table_name):
    cursor = connection.cursor()
    
    # Obtener ultimo Id de la tabla especificada
    query = f"SELECT MAX(Id) FROM {table_name.capitalize()}"
    cursor.execute(query)
    max_id = cursor.fetchone()[0]

    # Calcular codigo bodega
    scope_identity = max_id + 1 if max_id else 1
    formatted_digit = f"{scope_identity:06}"

    # Utiliza el nombre de la tabla para determinar el prefijo del código
    prefix = "BE" if table_name.lower() == "equipo" else "BC"
    
    code = f"{prefix}{formatted_digit}"
    return code


def if_default(text, default):
    if text is None:
        return default
    if isinstance(text, (str, int)):
        cleaned = str(text).strip()
        cleaned = re.sub(r'\s+', ' ', cleaned)
        if cleaned in ['*','-']:
            return default
        elif cleaned != '':
            return cleaned
        else:
            default
    return None

def process_equipo():
    wb = openpyxl.load_workbook(filename)
    sheet = wb[('inventario solo equipo').upper()]
    headers = [cell.value for cell in sheet[1]]

    # Encabezados esperados
    expected_headers = ['CATEGORIA', 'ESTADO', 'DESCRIPCION', 'MARCA', 'MODELO', 'ACTIVOTEC',  'SERIE', 'OBSERVACIONES', 'CONDICION', 'ESTANTE', 'CANTIDAD']
    print("hi")
    # Asegurarse de que todos los encabezados esperados estén presentes
    if all(header in headers for header in expected_headers):
        print("[...] Equipo")

        i = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Asigna los valores a variables según los encabezados dinámicamente, debe ser el mismo orden que el de expected headers
            categoria, estado, descripcion, marca, modelo, activoTec, serie, observaciones, condicion, estante, cantidad = (
                row[headers.index(header)] for header in expected_headers
            )

            # Verifica si todas las celdas de la fila están vacías o nulas
            if all(value is None or str(value).strip() == '' for value in [categoria, estado, descripcion, marca, modelo, activoTec, serie, observaciones, condicion, estante, cantidad ]):
                print("Linea vacia")
                break
            
            i+=1
            print ("Linea:",i)
    
            # Formateo
            categoriaId = getCategoria("equipo", categoria.upper())
            estadoId = estado_dict.get(estado.upper(), 1)
            fechaRegistro = datetime.now(pytz.timezone('America/Costa_Rica'))
            descripcion = if_default(descripcion, 'SIN DEFINIR')
            marca = if_default(marca, None)
            modelo = if_default(modelo, None)
            activoTec = 'SIN DEFINIR' if activoTec is None else str(activoTec).replace(" ", "").upper()
            serie = if_default(serie, None)
            observaciones = if_default(observaciones, None)
            condicion = condicion_dict.get(condicion.upper(), 1) # Condicion: Convierte la condición a mayúsculas y asigna por defecto "BUENO" si no es válida
            estante =  if_default(None if estante is None else estante.upper(), "N/A")
            
            # Insertar a DB
            cursor = connection.cursor()
            for _ in range(int(cantidad)):
                activoBodega = getActivoBodega("equipo") # Distinto para cada entidad
                # Id, CategoriaId, EstadoId, FechaRegistro, Descripcion, Marca, Modelo, ActivoBodega, ActivoTec, Serie, Observaciones, Condicion, Estante
                values = (categoriaId, estadoId, fechaRegistro, descripcion, marca, modelo, activoBodega, activoTec, serie, observaciones, condicion, estante)
                query = """
                    INSERT INTO Equipo (CategoriaId, EstadoId, FechaRegistro, Descripcion, Marca, Modelo, ActivoBodega, ActivoTec, Serie, Observaciones, Condicion, Estante)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, values)
                connection.commit()

    print("[DONE] Equipo")

def process_componentes():
    wb = openpyxl.load_workbook(filename)
    sheet = wb[('inventario componentes').upper()]
    headers = [cell.value for cell in sheet[1]]

    # Encabezados esperados
    expected_headers = ['CATEGORIA', 'DESCRIPCION', 'NO. PARTE', 'CANTIDAD', 'OBSERVACIONES', 'CONDICION', 'ESTANTE']
    i = 1
    # Asegurarse de que todos los encabezados esperados estén presentes
    if all(header in headers for header in expected_headers):
        print("[...] Componentes")
            
        for row in sheet.iter_rows(min_row=2):

            
            # Asigna los valores a variables según los encabezados dinámicamente, debe ser el mismo orden que el de expected headers
            categoria, descripcion, no_parte, cantidad, observaciones, condicion, estante = (
                row[headers.index(header)].value for header in expected_headers
            )
            
            # Verifica si todas las celdas de la fila están vacías o nulas
            if all(value is None or str(value).strip() == '' for value in [categoria, descripcion, no_parte, cantidad, observaciones, condicion, estante]):
                print("Linea vacia")
                break
            
            i+=1
            print ("Linea:",i)            

            # Formateo
            categoriaId = getCategoria("componente", categoria.upper())
            estadoId = 1 if (cantidad is not None and int(cantidad) > 0) else 3  # 1:Disponible - 3:Agotado
            fechaRegistro = datetime.now(pytz.timezone('America/Costa_Rica'))
            descripcion = if_default(descripcion, 'SIN DEFINIR')
            cantidadTotal = int(cantidad) if cantidad else 0     # Valor defecto si no esta, es cero
            cantidadDisponible = cantidadTotal
            activoBodega = getActivoBodega("componente")
            observaciones = if_default(observaciones, None)
            estante =  if_default(None if estante is None else estante.upper(), "SIN DEFINIR")
            condicion = condicion_dict.get(condicion.upper(), 1) if condicion else 1 # Condicion: Convierte la condición a mayúsculas y asigna por defecto "BUENO" si no es válida
            no_parte = if_default(no_parte, "N/A")
            
            # Insertar a dB
            cursor = connection.cursor()
            # Id, CategoriaId, EstadoId, FechaRegistro, Descripcion, Cantidad, CantidadDisponible, ActivoBodega, Observaciones, Estante, Condicion, Modelo
            values = (categoriaId, estadoId, fechaRegistro, descripcion, cantidadTotal, cantidadDisponible, activoBodega, observaciones, estante, condicion, no_parte,)
            query = """
                INSERT INTO Componente (CategoriaId, EstadoId, FechaRegistro, Descripcion, CantidadTotal, CantidadDisponible, ActivoBodega, Observaciones, Estante, Condicion, NoParte)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(query, values)
            connection.commit()
    
    print("[DONE] Componentes")



# Ejecuta la función main al abrir el archivo
if __name__ == "__main__":
    # Configura la conexión a la base de datos
    connection = mysql.connector.connect(
        host="localhost",
        user="root",
        password="admin",
        database="sibe_db"
    )
    filename = 'Inventario_1S24.xlsx'

    # Procesar
    process_equipo()
    process_componentes()

    # Cierra la conexión a la base de datos
    connection.close()